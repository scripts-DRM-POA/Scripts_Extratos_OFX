import pandas as pd
import re
import os
import pdfplumber
import itertools
from decimal import Decimal, ROUND_HALF_UP

# =====================================================
# CONFIGURAÇÃO
# =====================================================

PASTA = r"C:\Users\michael.lasch\Desktop\Repositorio_Padrao_Para_Rodar_Codigos"
ARQUIVO_PDF = "extrato_unicred.pdf"
ARQUIVO_SAIDA = "Extrato_Unicred_Auditoria_V5.xlsx"

CAMINHO_PDF = os.path.join(PASTA, ARQUIVO_PDF)
CAMINHO_SAIDA = os.path.join(PASTA, ARQUIVO_SAIDA)

TOLERANCIA = Decimal("0.01")

# =====================================================
# FUNÇÕES AUXILIARES
# =====================================================

def br_to_decimal(valor_str):
    valor_str = valor_str.replace('.', '').replace(',', '.')
    return Decimal(valor_str).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)

def extrair_valores_linha(linha):
    padrao = r'(-?\d{1,3}(?:\.\d{3})*,\d{2})'
    return re.findall(padrao, linha)

# =====================================================
# EXTRAÇÃO DO PDF
# =====================================================

def extrair_texto_pdf(caminho):
    texto_completo = ""
    with pdfplumber.open(caminho) as pdf:
        for pagina in pdf.pages:
            texto = pagina.extract_text()
            if texto:
                texto_completo += texto + "\n"
    return texto_completo

# =====================================================
# PARSER
# =====================================================

def processar_extrato(texto):
    
    linhas = texto.splitlines()
    registros = []
    saldo_anterior = None
    
    for linha in linhas:
        
        if "Saldo Anterior" in linha:
            valores = extrair_valores_linha(linha)
            if valores:
                saldo_anterior = br_to_decimal(valores[-1])
            continue
        
        if re.match(r"\d{2}/\d{2}/\d{4}", linha):
            
            data = linha[:10]
            valores = extrair_valores_linha(linha)
            
            if not valores:
                continue
            
            valor_mov = br_to_decimal(valores[0])
            saldo_info = br_to_decimal(valores[1]) if len(valores) > 1 else None
            
            historico = linha[11:]
            
            registros.append({
                "Data": data,
                "Historico": historico,
                "Valor": valor_mov,
                "Saldo_Informado": saldo_info
            })
    
    df = pd.DataFrame(registros).reset_index(drop=True)
    return df, saldo_anterior

# =====================================================
# RESOLVER BLOCO ENTRE SALDOS
# =====================================================

def resolver_bloco(bloco, saldo_anterior, saldo_final):
    
    bloco = bloco.copy()
    bloco["Tipo"] = None
    
    indices_livres = []
    
    # Pré-classificação heurística
    for i, row in bloco.iterrows():
        
        historico = row["Historico"].upper()
        
        if "INTEGR PARC CAPITAL" in historico:
            bloco.at[i, "Tipo"] = "D"
            
        elif "RECEB" in historico:
            bloco.at[i, "Tipo"] = "C"
            
        else:
            indices_livres.append(i)
    
    # Saldo base com os já classificados
    saldo_base = saldo_anterior
    
    for i, row in bloco.iterrows():
        if row["Tipo"] == "C":
            saldo_base += row["Valor"]
        elif row["Tipo"] == "D":
            saldo_base -= row["Valor"]
    
    # Testar combinações
    n = len(indices_livres)
    
    for combinacao in itertools.product(["C","D"], repeat=n):
        
        saldo_teste = saldo_base
        
        for idx, tipo in zip(indices_livres, combinacao):
            valor = bloco.at[idx, "Valor"]
            if tipo == "C":
                saldo_teste += valor
            else:
                saldo_teste -= valor
        
        if abs(saldo_teste - saldo_final) <= TOLERANCIA:
            
            for idx, tipo in zip(indices_livres, combinacao):
                bloco.at[idx, "Tipo"] = tipo
            
            return bloco, True
    
    return bloco, False

# =====================================================
# CLASSIFICAÇÃO POR INTERVALO
# =====================================================

def classificar_por_intervalo(df, saldo_inicial):
    
    df["Tipo"] = None
    saldo_confirmado = saldo_inicial
    inicio_bloco = 0
    logs = []
    
    for i in range(len(df)):
        
        if pd.notna(df.loc[i, "Saldo_Informado"]):
            
            saldo_final = df.loc[i, "Saldo_Informado"]
            
            bloco = df.loc[inicio_bloco:i].copy()
            
            bloco_resolvido, ok = resolver_bloco(
                bloco,
                saldo_confirmado,
                saldo_final
            )
            
            if ok:
                df.loc[inicio_bloco:i, "Tipo"] = bloco_resolvido["Tipo"]
            else:
                logs.append(
                    f"Bloco {inicio_bloco}-{i} não fechou | "
                    f"Saldo anterior: {saldo_confirmado} | "
                    f"Saldo final: {saldo_final}"
                )
            
            saldo_confirmado = saldo_final
            inicio_bloco = i + 1
    
    return df, logs

# =====================================================
# SALDO LINHA A LINHA
# =====================================================

def calcular_saldo_linha_a_linha(df, saldo_inicial):
    
    saldo_corrente = saldo_inicial
    saldos = []
    
    for _, row in df.iterrows():
        
        tipo = row["Tipo"]
        valor = row["Valor"]
        
        if tipo == "C":
            saldo_corrente += valor
        elif tipo == "D":
            saldo_corrente -= valor
        
        saldos.append(saldo_corrente)
    
    df["Saldo_Recalculado"] = saldos
    return df

# =====================================================
# EXECUÇÃO
# =====================================================

print("📄 Lendo PDF...")
texto = extrair_texto_pdf(CAMINHO_PDF)

print("🔎 Processando extrato...")
df, saldo_inicial = processar_extrato(texto)

print("🧮 Classificando por intervalo...")
df, logs = classificar_por_intervalo(df, saldo_inicial)

print("📊 Calculando saldo linha a linha...")
df = calcular_saldo_linha_a_linha(df, saldo_inicial)

df["Diferenca_Saldo"] = (
    df["Saldo_Informado"] - df["Saldo_Recalculado"]
)

inconsistencias = df[
    (df["Diferenca_Saldo"].notna()) &
    (abs(df["Diferenca_Saldo"]) > TOLERANCIA)
]

print("💾 Gerando Excel...")

colunas_monetarias = [
    "Valor",
    "Saldo_Informado",
    "Saldo_Recalculado"
]

# 🔹 Apenas garantir tipo numérico (SEM replace)
for col in colunas_monetarias:
    if col in df.columns:
        df[col] = pd.to_numeric(df[col], errors="coerce")

    if col in inconsistencias.columns:
        inconsistencias[col] = pd.to_numeric(inconsistencias[col], errors="coerce")

# ==========================
# EXPORTAÇÃO
# ==========================

with pd.ExcelWriter(CAMINHO_SAIDA, engine="openpyxl") as writer:
    df.to_excel(writer, sheet_name="Extrato_Processado", index=False)
    inconsistencias.to_excel(writer, sheet_name="Inconsistencias", index=False)
    pd.DataFrame(logs, columns=["Log"]).to_excel(
        writer, sheet_name="Log_Bloco", index=False
    )

# ==========================
# APLICAR SOMENTE FORMATAÇÃO VISUAL
# ==========================

from openpyxl import load_workbook

wb = load_workbook(CAMINHO_SAIDA)

for aba in ["Extrato_Processado", "Inconsistencias"]:

    ws = wb[aba]
    headers = [cell.value for cell in ws[1]]

    for col_nome in colunas_monetarias:
        if col_nome in headers:
            col_idx = headers.index(col_nome) + 1

            for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = '#,##0.00'

wb.save(CAMINHO_SAIDA)

print("✅ Versão final estável aplicada com sucesso.")
print(f"Arquivo salvo em: {CAMINHO_SAIDA}")