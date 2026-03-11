import fitz, re, pandas as pd, os
from datetime import datetime
from tkinter import Tk, filedialog
from openpyxl.utils import get_column_letter
from openpyxl.styles import numbers

def process_pdf(pdf_path):
    doc = fitz.open(pdf_path)

    def ordered_text_all_pages():
        out = []
        for page in doc:
            blocks = page.get_text("blocks")
            blocks.sort(key=lambda b: (round(b[1],2), round(b[0],2)))
            for b in blocks:
                t = b[4].replace("\xa0"," ").rstrip()
                if t:
                    out.append(t)
        return "\n".join(out)

    raw = ordered_text_all_pages()
    start = raw.find("Movimentação")
    end = raw.find("Saldos por Período", start if start!=-1 else 0)
    if start == -1: start = 0
    if end == -1: end = len(raw)
    mov = raw[start:end]

    lines = [re.sub(r"\s{2,}", " ", l.strip()) for l in mov.splitlines()]
    lines = [l for l in lines if l and not l.startswith("Pagina:")]

    date_re = re.compile(r"^(\d{2})/(\d{2})(?:/(\d{4}))?$")
    current_date = None
    current_year = 2023
    current_month = None
    pending_desc_parts = []
    rows = []

    skip_line_patterns = (
        "saldo do dia", "saldo em", "subtotal", "saldos por", "contamax empresarial",
        "cdb contamax", "movimentação mensal", "posição consolidada"
    )

    def flush_desc():
        txt = " ".join(pending_desc_parts).strip(" -")
        return re.sub(r"\s{2,}", " ", txt)

    for line in lines:
        low = line.lower()
        if any(pat in low for pat in skip_line_patterns):
            continue

        md = date_re.match(line.split()[0]) if line else None
        if md:
            d, m, y = md.group(1), md.group(2), md.group(3) or "2023"
            if len(y) == 2: y = "20"+y
            current_date = f"{d}/{m}/{y}"
            current_year = int(y)
            current_month = int(m)
            rest = line[len(md.group(0)):].strip(" -")
            rest = re.sub(r"(\d{1,3}(?:\.\d{3})*,\d{2}-?)\s*(\d{1,3}(?:\.\d{3})*,\d{2}-?)?\s*$", "", rest).strip(" -")
            pending_desc_parts = [rest] if rest else []
            tokens = list(re.finditer(r"\d{1,3}(?:\.\d{3})*,\d{2}-?", line))
            tokens = [t.group(0) for t in tokens if t.group(0) != "0,00"]
            if tokens:
                tok = tokens[0]
                val = float(tok.replace(".","").replace(",", ".").replace("-", ""))
                if tok.endswith("-"):
                    val = -val
                desc = flush_desc()
                if not desc:
                    desc = rest
                if desc:
                    rows.append([current_date, current_year, current_month, desc, val, "C" if val>0 else "D"])
            continue

        if current_date is None:
            continue

        money_tokens = list(re.finditer(r"\d{1,3}(?:\.\d{3})*,\d{2}-?", line))
        if not money_tokens:
            if not re.fullmatch(r"\d{4,}", line):
                pending_desc_parts.append(line)
            continue

        tokens = [t.group(0) for t in money_tokens if t.group(0) != "0,00"]
        if not tokens:
            continue
        tok = tokens[0]
        val = float(tok.replace(".","").replace(",", ".").replace("-", ""))
        if tok.endswith("-"):
            val = -val
        desc_inline = re.sub(r"\s*(\d{1,3}(?:\.\d{3})*,\d{2}-?)(\s+\d{1,3}(?:\.\d{3})*,\d{2}-?)?\s*$", "", line).strip(" -")
        if desc_inline and desc_inline not in pending_desc_parts:
            pending_desc_parts.append(desc_inline)
        desc = flush_desc()
        if not desc:
            desc = desc_inline
        if not desc:
            continue
        rows.append([current_date, current_year, current_month, desc, val, "C" if val>0 else "D"])
        pending_desc_parts = []

    df = pd.DataFrame(rows, columns=["Data", "Ano", "Mês", "Descrição da operação", "Valor", "Tipo"])

    def valid_date(s):
        try:
            datetime.strptime(s, "%d/%m/%Y")
            return True
        except:
            return False

    df = df[df["Data"].apply(valid_date)].copy()
    df["Descrição da operação"] = df["Descrição da operação"].str.replace(r"\s{2,}", " ", regex=True).str.strip()
    df["__dt"] = pd.to_datetime(df["Data"], format="%d/%m/%Y")
    df["__seq"] = range(len(df))
    df.sort_values(["__dt","__seq"], inplace=True)
    df.drop(columns=["__dt","__seq"], inplace=True)

    return df

# Selecionar múltiplos PDFs
Tk().withdraw()
pdf_files = filedialog.askopenfilenames(title="Selecione os arquivos PDF", filetypes=[("PDF files", "*.pdf")])
if not pdf_files:
    raise SystemExit("Nenhum arquivo selecionado.")

# Criar Excel consolidado
output_file = os.path.join(os.path.dirname(pdf_files[0]), "consolidado_lancamentos.xlsx")
with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
    consolidated = []
    for pdf_path in pdf_files:
        df = process_pdf(pdf_path)
        sheet_name = os.path.splitext(os.path.basename(pdf_path))[0][:31]  # Nome da aba <=31 chars
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        consolidated.append(df)
        
        # Formatação de colunas
        ws = writer.sheets[sheet_name]
        col_widths = [12, 6, 5, 70, 16, 5]  # A,B,C,D,E,F
        for i, width in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = width
        # Formato monetário na coluna E (Valor)
        for cell in ws['E'][1:]:
            cell.number_format = 'R$ #,##0.00;[Red]-R$ #,##0.00'

    # Aba consolidada
    if consolidated:
        df_consol = pd.concat(consolidated, ignore_index=True)
        df_consol.to_excel(writer, index=False, sheet_name="Consolidado")
        ws = writer.sheets["Consolidado"]
        col_widths = [12, 6, 5, 70, 16, 5]
        for i, width in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = width
        for cell in ws['E'][1:]:
            cell.number_format = 'R$ #,##0.00;[Red]-R$ #,##0.00'

print(f"Arquivo consolidado salvo em: {output_file}")
